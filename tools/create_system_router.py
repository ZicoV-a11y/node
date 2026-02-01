"""
Create System Router Excel Workbook
Generates an Excel file matching the AV routing interface style
with devices, inputs, outputs, and color coding
"""

import os
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

def create_system_router(output_path=None):
    if output_path is None:
        output_path = os.path.join(os.path.dirname(__file__), '..', '.tmp', 'system_router.xlsx')

    wb = Workbook()

    # Colors matching the UI
    DARK_BG = "1a1a2e"
    HEADER_BG = "2d2d44"
    ROW_BG = "252538"
    ACCENT_BLUE = "4472C4"

    # Color swatches from the UI
    COLORS = {
        "Red": "FF0000",
        "Green": "00FF00",
        "Blue": "0000FF",
        "Cyan": "00FFFF",
        "Magenta": "FF00FF",
        "Purple": "8B00FF",
        "Yellow": "FFFF00",
        "Orange": "FFA500"
    }

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    subheader_font = Font(bold=True, color="AAAAAA", size=10)
    cell_font = Font(color="FFFFFF", size=10)
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    row_fill = PatternFill(start_color=ROW_BG, end_color=ROW_BG, fill_type="solid")
    dark_fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color="444466"),
        right=Side(style='thin', color="444466"),
        top=Side(style='thin', color="444466"),
        bottom=Side(style='thin', color="444466")
    )

    # ========== DEVICES SHEET ==========
    ws_devices = wb.active
    ws_devices.title = "Devices"
    ws_devices.sheet_properties.tabColor = "4472C4"

    # Apply dark background to used area
    for row in range(1, 30):
        for col in range(1, 12):
            ws_devices.cell(row=row, column=col).fill = dark_fill

    # Device header
    ws_devices.merge_cells('B2:H2')
    device_title = ws_devices['B2']
    device_title.value = "Laptop"
    device_title.font = Font(bold=True, color="FFFFFF", size=14)
    device_title.fill = header_fill
    device_title.alignment = Alignment(horizontal='center', vertical='center')

    # Color row
    color_list = list(COLORS.items())
    for i, (name, hex_color) in enumerate(color_list):
        cell = ws_devices.cell(row=3, column=2+i)
        cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        cell.border = thin_border
        ws_devices.column_dimensions[get_column_letter(2+i)].width = 4

    # SYSTEMS label
    ws_devices.merge_cells('B4:H4')
    systems_label = ws_devices['B4']
    systems_label.value = "SYSTEMS"
    systems_label.font = subheader_font
    systems_label.fill = header_fill
    systems_label.alignment = Alignment(horizontal='center')

    # INPUT section header
    ws_devices['B6'] = "INPUT"
    ws_devices['B6'].font = header_font
    ws_devices['B6'].fill = header_fill
    ws_devices['C6'] = "+"
    ws_devices['C6'].font = header_font
    ws_devices['C6'].fill = header_fill
    ws_devices['C6'].alignment = Alignment(horizontal='right')

    # INPUT column headers
    ws_devices['B7'] = "SOURCE"
    ws_devices['C7'] = "NAME"
    ws_devices['B7'].font = subheader_font
    ws_devices['C7'].font = subheader_font
    ws_devices['B7'].fill = row_fill
    ws_devices['C7'].fill = row_fill

    # OUTPUT section header
    ws_devices['E6'] = "OUTPUT"
    ws_devices['E6'].font = header_font
    ws_devices['E6'].fill = header_fill
    ws_devices.merge_cells('E6:G6')
    ws_devices['H6'] = "+"
    ws_devices['H6'].font = header_font
    ws_devices['H6'].fill = header_fill
    ws_devices['H6'].alignment = Alignment(horizontal='right')

    # OUTPUT column headers
    ws_devices['E7'] = "SOURCE"
    ws_devices['F7'] = "NAME"
    ws_devices['E7'].font = subheader_font
    ws_devices['F7'].font = subheader_font
    ws_devices['E7'].fill = row_fill
    ws_devices['F7'].fill = row_fill

    # OUTPUT data rows
    output_data = [
        ("HDMI", "HDMI"),
        ("SDI", "USB-C"),
    ]

    for i, (source, name) in enumerate(output_data, 8):
        # Source dropdown cell
        source_cell = ws_devices.cell(row=i, column=5, value=source)
        source_cell.font = cell_font
        source_cell.fill = row_fill
        source_cell.border = thin_border

        # Name cell
        name_cell = ws_devices.cell(row=i, column=6, value=name)
        name_cell.font = cell_font
        name_cell.fill = row_fill
        name_cell.border = thin_border

    # Add data validation for source dropdown
    source_dv = DataValidation(
        type="list",
        formula1='"HDMI,SDI,DisplayPort,USB-C,Thunderbolt,VGA,DVI,Composite"',
        allow_blank=True
    )
    ws_devices.add_data_validation(source_dv)
    source_dv.add('E8:E20')

    # Column widths
    ws_devices.column_dimensions['A'].width = 3
    ws_devices.column_dimensions['B'].width = 15
    ws_devices.column_dimensions['C'].width = 15
    ws_devices.column_dimensions['D'].width = 3
    ws_devices.column_dimensions['E'].width = 15
    ws_devices.column_dimensions['F'].width = 15
    ws_devices.column_dimensions['G'].width = 5
    ws_devices.column_dimensions['H'].width = 5

    # ========== SOURCES SHEET (lookup table) ==========
    ws_sources = wb.create_sheet("Sources")
    ws_sources.sheet_properties.tabColor = "00AA00"

    source_headers = ["SourceID", "Source Type", "Connector", "Max Resolution", "Audio Support"]
    for col, header in enumerate(source_headers, 1):
        cell = ws_sources.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    source_data = [
        [1, "HDMI", "HDMI 2.1", "8K@60Hz", "Yes"],
        [2, "SDI", "BNC", "4K@60Hz", "Embedded"],
        [3, "DisplayPort", "DP 1.4", "8K@60Hz", "Yes"],
        [4, "USB-C", "USB-C", "4K@60Hz", "Yes"],
        [5, "Thunderbolt", "TB4", "8K@60Hz", "Yes"],
        [6, "VGA", "DE-15", "1080p@60Hz", "No"],
        [7, "DVI", "DVI-D", "2560x1600", "No"],
        [8, "Composite", "RCA", "480i", "Separate"],
    ]
    for row_idx, row_data in enumerate(source_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_sources.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    for i, width in enumerate([10, 14, 12, 14, 14], 1):
        ws_sources.column_dimensions[get_column_letter(i)].width = width

    # Named range for source types
    source_range = DefinedName('SourceTypes', attr_text='Sources!$B$2:$B$20')
    wb.defined_names.add(source_range)

    # ========== COLORS SHEET ==========
    ws_colors = wb.create_sheet("Colors")
    ws_colors.sheet_properties.tabColor = "FF0000"

    color_headers = ["ColorID", "Color Name", "Hex Code", "Use For"]
    for col, header in enumerate(color_headers, 1):
        cell = ws_colors.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, (name, hex_code) in enumerate(COLORS.items(), 2):
        ws_colors.cell(row=row_idx, column=1, value=row_idx-1)
        ws_colors.cell(row=row_idx, column=2, value=name)
        ws_colors.cell(row=row_idx, column=3, value=f"#{hex_code}")
        # Color preview cell
        preview = ws_colors.cell(row=row_idx, column=4, value="")
        preview.fill = PatternFill(start_color=hex_code, end_color=hex_code, fill_type="solid")
        for col in range(1, 5):
            ws_colors.cell(row=row_idx, column=col).border = thin_border

    for i, width in enumerate([10, 14, 12, 12], 1):
        ws_colors.column_dimensions[get_column_letter(i)].width = width

    # ========== SAVE ==========
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"System router created: {output_path}")
    return output_path

if __name__ == "__main__":
    create_system_router()
