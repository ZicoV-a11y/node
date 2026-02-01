"""
Create Project Management Excel Workbook
Generates an Excel file with connected tables: People, Projects, Tasks, Dashboard
Includes: dropdown validation, lookups, and calculations
"""

import os
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

def create_project_tracker(output_path=None):
    if output_path is None:
        output_path = os.path.join(os.path.dirname(__file__), '..', '.tmp', 'project_tracker.xlsx')

    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ========== PEOPLE SHEET ==========
    ws_people = wb.active
    ws_people.title = "People"

    people_headers = ["PersonID", "Name", "Role", "Email", "Hourly Rate"]
    for col, header in enumerate(people_headers, 1):
        cell = ws_people.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Sample data
    people_data = [
        [1, "Alice Johnson", "Developer", "alice@example.com", 75],
        [2, "Bob Smith", "Designer", "bob@example.com", 65],
        [3, "Carol Williams", "Manager", "carol@example.com", 85],
        [4, "David Brown", "Developer", "david@example.com", 70],
    ]
    for row_idx, row_data in enumerate(people_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_people.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Set column widths
    ws_people.column_dimensions['A'].width = 10
    ws_people.column_dimensions['B'].width = 18
    ws_people.column_dimensions['C'].width = 12
    ws_people.column_dimensions['D'].width = 22
    ws_people.column_dimensions['E'].width = 12

    # Create named range for People names (for dropdowns)
    people_range = DefinedName('PeopleNames', attr_text='People!$B$2:$B$100')
    wb.defined_names.add(people_range)

    # ========== PROJECTS SHEET ==========
    ws_projects = wb.create_sheet("Projects")

    project_headers = ["ProjectID", "Project Name", "Client", "Start Date", "End Date", "Budget"]
    for col, header in enumerate(project_headers, 1):
        cell = ws_projects.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    project_data = [
        [1, "Website Redesign", "Acme Corp", "2024-01-15", "2024-04-30", 25000],
        [2, "Mobile App", "TechStart Inc", "2024-02-01", "2024-06-30", 45000],
        [3, "Database Migration", "Global Ltd", "2024-03-01", "2024-05-15", 18000],
    ]
    for row_idx, row_data in enumerate(project_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_projects.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    ws_projects.column_dimensions['A'].width = 10
    ws_projects.column_dimensions['B'].width = 20
    ws_projects.column_dimensions['C'].width = 15
    ws_projects.column_dimensions['D'].width = 12
    ws_projects.column_dimensions['E'].width = 12
    ws_projects.column_dimensions['F'].width = 12

    # Named range for Project names
    project_range = DefinedName('ProjectNames', attr_text='Projects!$B$2:$B$100')
    wb.defined_names.add(project_range)

    # ========== TASKS SHEET ==========
    ws_tasks = wb.create_sheet("Tasks")

    task_headers = ["TaskID", "Task Name", "Project", "Assignee", "Status", "Priority", "Hours Est.", "Hours Actual", "Due Date"]
    for col, header in enumerate(task_headers, 1):
        cell = ws_tasks.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    task_data = [
        [1, "Design mockups", "Website Redesign", "Bob Smith", "Completed", "High", 20, 18, "2024-02-01"],
        [2, "Frontend development", "Website Redesign", "Alice Johnson", "In Progress", "High", 40, 25, "2024-03-15"],
        [3, "API integration", "Mobile App", "David Brown", "Not Started", "Medium", 30, 0, "2024-04-01"],
        [4, "User testing", "Website Redesign", "Carol Williams", "Not Started", "Medium", 15, 0, "2024-04-15"],
        [5, "Database schema", "Database Migration", "Alice Johnson", "In Progress", "High", 25, 12, "2024-03-20"],
    ]
    for row_idx, row_data in enumerate(task_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_tasks.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Column widths
    col_widths = [8, 20, 18, 15, 12, 10, 12, 12, 12]
    for i, width in enumerate(col_widths, 1):
        ws_tasks.column_dimensions[get_column_letter(i)].width = width

    # Data Validation: Status dropdown
    status_dv = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,On Hold,Completed,Cancelled"',
        allow_blank=True
    )
    status_dv.error = "Please select a valid status"
    status_dv.errorTitle = "Invalid Status"
    ws_tasks.add_data_validation(status_dv)
    status_dv.add('E2:E100')

    # Data Validation: Priority dropdown
    priority_dv = DataValidation(
        type="list",
        formula1='"Low,Medium,High,Critical"',
        allow_blank=True
    )
    ws_tasks.add_data_validation(priority_dv)
    priority_dv.add('F2:F100')

    # Data Validation: Project dropdown (from Projects sheet)
    project_dv = DataValidation(
        type="list",
        formula1='=ProjectNames',
        allow_blank=True
    )
    project_dv.error = "Please select a project from the list"
    ws_tasks.add_data_validation(project_dv)
    project_dv.add('C2:C100')

    # Data Validation: Assignee dropdown (from People sheet)
    assignee_dv = DataValidation(
        type="list",
        formula1='=PeopleNames',
        allow_blank=True
    )
    assignee_dv.error = "Please select a person from the list"
    ws_tasks.add_data_validation(assignee_dv)
    assignee_dv.add('D2:D100')

    # ========== DASHBOARD SHEET ==========
    ws_dash = wb.create_sheet("Dashboard")

    # Title
    ws_dash['A1'] = "PROJECT DASHBOARD"
    ws_dash['A1'].font = Font(bold=True, size=16)
    ws_dash.merge_cells('A1:D1')

    # Summary section
    ws_dash['A3'] = "SUMMARY METRICS"
    ws_dash['A3'].font = Font(bold=True, size=12)

    summary_labels = [
        ("Total Tasks:", '=COUNTA(Tasks!A2:A100)'),
        ("Completed:", '=COUNTIF(Tasks!E:E,"Completed")'),
        ("In Progress:", '=COUNTIF(Tasks!E:E,"In Progress")'),
        ("Not Started:", '=COUNTIF(Tasks!E:E,"Not Started")'),
        ("Completion Rate:", '=IF(COUNTA(Tasks!A2:A100)>0,COUNTIF(Tasks!E:E,"Completed")/COUNTA(Tasks!A2:A100),0)'),
        ("Total Hours Est.:", '=SUM(Tasks!G:G)'),
        ("Total Hours Actual:", '=SUM(Tasks!H:H)'),
    ]

    for i, (label, formula) in enumerate(summary_labels, 4):
        ws_dash.cell(row=i, column=1, value=label).font = Font(bold=True)
        cell = ws_dash.cell(row=i, column=2, value=formula)
        if "Rate" in label:
            cell.number_format = '0%'

    # Tasks by Project section
    ws_dash['A13'] = "TASKS BY PROJECT"
    ws_dash['A13'].font = Font(bold=True, size=12)

    ws_dash['A14'] = "Project"
    ws_dash['B14'] = "Total"
    ws_dash['C14'] = "Completed"
    ws_dash['D14'] = "% Done"
    for col in range(1, 5):
        ws_dash.cell(row=14, column=col).font = header_font
        ws_dash.cell(row=14, column=col).fill = header_fill

    # Project summary formulas (using COUNTIF)
    for i, project in enumerate(["Website Redesign", "Mobile App", "Database Migration"], 15):
        ws_dash.cell(row=i, column=1, value=project)
        ws_dash.cell(row=i, column=2, value=f'=COUNTIF(Tasks!C:C,"{project}")')
        ws_dash.cell(row=i, column=3, value=f'=COUNTIFS(Tasks!C:C,"{project}",Tasks!E:E,"Completed")')
        pct_cell = ws_dash.cell(row=i, column=4, value=f'=IF(B{i}>0,C{i}/B{i},0)')
        pct_cell.number_format = '0%'

    # Tasks by Person section
    ws_dash['A20'] = "TASKS BY ASSIGNEE"
    ws_dash['A20'].font = Font(bold=True, size=12)

    ws_dash['A21'] = "Assignee"
    ws_dash['B21'] = "Assigned"
    ws_dash['C21'] = "Completed"
    ws_dash['D21'] = "Hours"
    for col in range(1, 5):
        ws_dash.cell(row=21, column=col).font = header_font
        ws_dash.cell(row=21, column=col).fill = header_fill

    # Assignee formulas
    for i, person in enumerate(["Alice Johnson", "Bob Smith", "Carol Williams", "David Brown"], 22):
        ws_dash.cell(row=i, column=1, value=person)
        ws_dash.cell(row=i, column=2, value=f'=COUNTIF(Tasks!D:D,"{person}")')
        ws_dash.cell(row=i, column=3, value=f'=COUNTIFS(Tasks!D:D,"{person}",Tasks!E:E,"Completed")')
        ws_dash.cell(row=i, column=4, value=f'=SUMIF(Tasks!D:D,"{person}",Tasks!H:H)')

    # Column widths for dashboard
    ws_dash.column_dimensions['A'].width = 20
    ws_dash.column_dimensions['B'].width = 12
    ws_dash.column_dimensions['C'].width = 12
    ws_dash.column_dimensions['D'].width = 12

    # ========== SAVE ==========
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Project tracker created: {output_path}")
    return output_path

if __name__ == "__main__":
    create_project_tracker()
