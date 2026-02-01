"""
AV Routing System with Interconnected Device Charts
Creates Excel workbook with:
- Master device list
- Individual device I/O tables
- Routing matrix (anchor points connecting devices)
- Signal flow tracking
"""

import os
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

def create_av_router(output_path=None):
    if output_path is None:
        output_path = os.path.join(os.path.dirname(__file__), '..', '.tmp', 'av_router.xlsx')

    wb = Workbook()

    # ===== STYLE DEFINITIONS =====
    COLORS = {
        "dark_bg": "1a1a2e",
        "header": "2d2d44",
        "row": "252538",
        "input_green": "2e7d32",
        "output_blue": "1565c0",
        "routing_purple": "6a1b9a",
        "accent": "4472C4"
    }

    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, color="FFFFFF", size=14)
    cell_font = Font(color="FFFFFF", size=10)
    link_font = Font(color="00BFFF", size=10, underline="single")

    def make_fill(color):
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    header_fill = make_fill(COLORS["header"])
    row_fill = make_fill(COLORS["row"])
    dark_fill = make_fill(COLORS["dark_bg"])
    input_fill = make_fill(COLORS["input_green"])
    output_fill = make_fill(COLORS["output_blue"])
    routing_fill = make_fill(COLORS["routing_purple"])

    thin_border = Border(
        left=Side(style='thin', color="444466"),
        right=Side(style='thin', color="444466"),
        top=Side(style='thin', color="444466"),
        bottom=Side(style='thin', color="444466")
    )

    # ===== DEVICE DEFINITIONS =====
    # Each device: (name, color_hex, inputs, outputs)
    DEVICES = [
        {
            "name": "Laptop",
            "color": "4472C4",
            "inputs": [],
            "outputs": [
                {"port": "HDMI", "name": "HDMI Out", "signal": "Video"},
                {"port": "USB-C", "name": "USB-C DP", "signal": "Video"},
                {"port": "3.5mm", "name": "Headphone", "signal": "Audio"},
            ]
        },
        {
            "name": "Video Switcher",
            "color": "7B68EE",
            "inputs": [
                {"port": "HDMI 1", "name": "Input 1", "signal": "Video"},
                {"port": "HDMI 2", "name": "Input 2", "signal": "Video"},
                {"port": "HDMI 3", "name": "Input 3", "signal": "Video"},
                {"port": "SDI 1", "name": "SDI In", "signal": "Video"},
            ],
            "outputs": [
                {"port": "HDMI", "name": "Program", "signal": "Video"},
                {"port": "SDI", "name": "SDI Out", "signal": "Video"},
            ]
        },
        {
            "name": "Display",
            "color": "20B2AA",
            "inputs": [
                {"port": "HDMI 1", "name": "HDMI Input", "signal": "Video"},
                {"port": "HDMI 2", "name": "HDMI 2", "signal": "Video"},
                {"port": "DP", "name": "DisplayPort", "signal": "Video"},
            ],
            "outputs": []
        },
        {
            "name": "Audio Mixer",
            "color": "FF6347",
            "inputs": [
                {"port": "XLR 1", "name": "Mic 1", "signal": "Audio"},
                {"port": "XLR 2", "name": "Mic 2", "signal": "Audio"},
                {"port": "Line 1", "name": "Line In", "signal": "Audio"},
                {"port": "USB", "name": "USB Audio", "signal": "Audio"},
            ],
            "outputs": [
                {"port": "XLR L", "name": "Main L", "signal": "Audio"},
                {"port": "XLR R", "name": "Main R", "signal": "Audio"},
                {"port": "USB", "name": "USB Out", "signal": "Audio"},
            ]
        },
    ]

    # ===== SHEET 1: MASTER DEVICE LIST =====
    ws_master = wb.active
    ws_master.title = "Devices"
    ws_master.sheet_properties.tabColor = COLORS["accent"]

    # Dark background
    for row in range(1, 40):
        for col in range(1, 15):
            ws_master.cell(row=row, column=col).fill = dark_fill

    # Title
    ws_master.merge_cells('B2:G2')
    ws_master['B2'] = "AV SYSTEM DEVICES"
    ws_master['B2'].font = title_font
    ws_master['B2'].fill = header_fill
    ws_master['B2'].alignment = Alignment(horizontal='center')

    # Headers
    master_headers = ["ID", "Device Name", "Type", "Inputs", "Outputs", "Sheet Link"]
    for col, header in enumerate(master_headers, 2):
        cell = ws_master.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Device rows
    for i, device in enumerate(DEVICES, 1):
        row = 4 + i
        ws_master.cell(row=row, column=2, value=i).font = cell_font
        ws_master.cell(row=row, column=2).fill = row_fill
        ws_master.cell(row=row, column=2).border = thin_border

        ws_master.cell(row=row, column=3, value=device["name"]).font = cell_font
        ws_master.cell(row=row, column=3).fill = make_fill(device["color"])
        ws_master.cell(row=row, column=3).border = thin_border

        # Determine type
        has_in = len(device["inputs"]) > 0
        has_out = len(device["outputs"]) > 0
        dev_type = "Source" if has_out and not has_in else "Destination" if has_in and not has_out else "Processor"
        ws_master.cell(row=row, column=4, value=dev_type).font = cell_font
        ws_master.cell(row=row, column=4).fill = row_fill
        ws_master.cell(row=row, column=4).border = thin_border

        ws_master.cell(row=row, column=5, value=len(device["inputs"])).font = cell_font
        ws_master.cell(row=row, column=5).fill = input_fill
        ws_master.cell(row=row, column=5).border = thin_border

        ws_master.cell(row=row, column=6, value=len(device["outputs"])).font = cell_font
        ws_master.cell(row=row, column=6).fill = output_fill
        ws_master.cell(row=row, column=6).border = thin_border

        # Hyperlink to device sheet
        link_cell = ws_master.cell(row=row, column=7, value=f"Go to {device['name']}")
        link_cell.font = link_font
        link_cell.fill = row_fill
        link_cell.border = thin_border
        link_cell.hyperlink = f"#'{device['name']}'!A1"

    # Column widths
    widths = [3, 6, 18, 12, 8, 8, 16]
    for i, w in enumerate(widths, 1):
        ws_master.column_dimensions[get_column_letter(i)].width = w

    # ===== CREATE DEVICE SHEETS =====
    all_inputs = []  # For routing dropdown
    all_outputs = []

    for device in DEVICES:
        ws = wb.create_sheet(device["name"])
        ws.sheet_properties.tabColor = device["color"]

        # Dark background
        for row in range(1, 25):
            for col in range(1, 12):
                ws.cell(row=row, column=col).fill = dark_fill

        # Device title bar
        ws.merge_cells('B2:I2')
        ws['B2'] = device["name"]
        ws['B2'].font = title_font
        ws['B2'].fill = make_fill(device["color"])
        ws['B2'].alignment = Alignment(horizontal='center', vertical='center')

        # Color indicator row
        colors_row = ["FF0000", "00FF00", "0000FF", "00FFFF", "FF00FF", "8B00FF", "FFFF00", "FFA500"]
        for i, c in enumerate(colors_row):
            cell = ws.cell(row=3, column=2+i)
            cell.fill = make_fill(c)
            cell.border = thin_border

        # ===== INPUTS SECTION =====
        ws.merge_cells('B5:D5')
        ws['B5'] = "INPUTS"
        ws['B5'].font = header_font
        ws['B5'].fill = input_fill
        ws['B5'].alignment = Alignment(horizontal='center')

        input_headers = ["Port", "Name", "Source →"]
        for col, header in enumerate(input_headers, 2):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = Font(bold=True, color="AAAAAA", size=9)
            cell.fill = row_fill
            cell.border = thin_border

        for i, inp in enumerate(device["inputs"], 7):
            ws.cell(row=i, column=2, value=inp["port"]).font = cell_font
            ws.cell(row=i, column=2).fill = row_fill
            ws.cell(row=i, column=2).border = thin_border

            ws.cell(row=i, column=3, value=inp["name"]).font = cell_font
            ws.cell(row=i, column=3).fill = row_fill
            ws.cell(row=i, column=3).border = thin_border

            # ANCHOR POINT: This is where another device's output connects
            anchor = ws.cell(row=i, column=4, value="[Select Source]")
            anchor.font = Font(color="00BFFF", size=10, italic=True)
            anchor.fill = make_fill("1a3a5c")
            anchor.border = Border(
                left=Side(style='medium', color="00BFFF"),
                right=Side(style='medium', color="00BFFF"),
                top=Side(style='medium', color="00BFFF"),
                bottom=Side(style='medium', color="00BFFF")
            )

            all_inputs.append(f"{device['name']}|{inp['port']}")

        # ===== OUTPUTS SECTION =====
        ws.merge_cells('F5:H5')
        ws['F5'] = "OUTPUTS"
        ws['F5'].font = header_font
        ws['F5'].fill = output_fill
        ws['F5'].alignment = Alignment(horizontal='center')

        output_headers = ["Port", "Name", "→ Destination"]
        for col, header in enumerate(output_headers, 6):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = Font(bold=True, color="AAAAAA", size=9)
            cell.fill = row_fill
            cell.border = thin_border

        for i, out in enumerate(device["outputs"], 7):
            ws.cell(row=i, column=6, value=out["port"]).font = cell_font
            ws.cell(row=i, column=6).fill = row_fill
            ws.cell(row=i, column=6).border = thin_border

            ws.cell(row=i, column=7, value=out["name"]).font = cell_font
            ws.cell(row=i, column=7).fill = row_fill
            ws.cell(row=i, column=7).border = thin_border

            # ANCHOR POINT: This is where this output connects to another device's input
            anchor = ws.cell(row=i, column=8, value="[Select Dest]")
            anchor.font = Font(color="FFA500", size=10, italic=True)
            anchor.fill = make_fill("3a2a1a")
            anchor.border = Border(
                left=Side(style='medium', color="FFA500"),
                right=Side(style='medium', color="FFA500"),
                top=Side(style='medium', color="FFA500"),
                bottom=Side(style='medium', color="FFA500")
            )

            all_outputs.append(f"{device['name']}|{out['port']}")

        # Back link
        ws['B20'] = "← Back to Devices"
        ws['B20'].font = link_font
        ws['B20'].hyperlink = "#'Devices'!A1"

        # Column widths
        for col, w in zip('ABCDEFGHI', [3, 12, 14, 20, 3, 12, 14, 20, 3]):
            ws.column_dimensions[col].width = w

    # ===== ROUTING MATRIX SHEET =====
    ws_routing = wb.create_sheet("Routing")
    ws_routing.sheet_properties.tabColor = COLORS["routing_purple"]

    # Dark background
    for row in range(1, 30):
        for col in range(1, 20):
            ws_routing.cell(row=row, column=col).fill = dark_fill

    # Title
    ws_routing.merge_cells('B2:H2')
    ws_routing['B2'] = "SIGNAL ROUTING MATRIX"
    ws_routing['B2'].font = title_font
    ws_routing['B2'].fill = routing_fill
    ws_routing['B2'].alignment = Alignment(horizontal='center')

    # Instructions
    ws_routing['B4'] = "Connect outputs to inputs below. Use dropdowns to select destinations."
    ws_routing['B4'].font = Font(color="AAAAAA", size=10, italic=True)

    # Headers
    routing_headers = ["#", "Source Device", "Output Port", "→", "Dest Device", "Input Port", "Signal Type", "Status"]
    for col, header in enumerate(routing_headers, 2):
        cell = ws_routing.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Pre-fill some routing rows
    routes = [
        (1, "Laptop", "HDMI Out", "→", "Video Switcher", "Input 1", "Video", "Active"),
        (2, "Video Switcher", "Program", "→", "Display", "HDMI Input", "Video", "Active"),
        (3, "Laptop", "3.5mm", "→", "Audio Mixer", "Line In", "Audio", "Active"),
        (4, "", "", "→", "", "", "", ""),
        (5, "", "", "→", "", "", "", ""),
    ]

    for i, route in enumerate(routes, 7):
        for j, val in enumerate(route):
            cell = ws_routing.cell(row=i, column=2+j, value=val)
            cell.font = cell_font
            cell.fill = row_fill
            cell.border = thin_border
            if j == 3:  # Arrow column
                cell.font = Font(color="00FF00", size=12, bold=True)
                cell.alignment = Alignment(horizontal='center')

    # Dropdowns for source device
    source_devices = ",".join([d["name"] for d in DEVICES if d["outputs"]])
    source_dv = DataValidation(type="list", formula1=f'"{source_devices}"', allow_blank=True)
    ws_routing.add_data_validation(source_dv)
    source_dv.add('C7:C20')

    # Dropdowns for dest device
    dest_devices = ",".join([d["name"] for d in DEVICES if d["inputs"]])
    dest_dv = DataValidation(type="list", formula1=f'"{dest_devices}"', allow_blank=True)
    ws_routing.add_data_validation(dest_dv)
    dest_dv.add('F7:F20')

    # Status dropdown
    status_dv = DataValidation(type="list", formula1='"Active,Inactive,Testing,Fault"', allow_blank=True)
    ws_routing.add_data_validation(status_dv)
    status_dv.add('I7:I20')

    # Column widths
    for col, w in zip('ABCDEFGHIJ', [3, 5, 16, 14, 5, 16, 14, 12, 10, 3]):
        ws_routing.column_dimensions[col].width = w

    # ===== ADD DROPDOWNS TO DEVICE SHEETS =====
    # Build output list for input dropdowns (source selection)
    output_list = []
    for d in DEVICES:
        for o in d["outputs"]:
            output_list.append(f"{d['name']}: {o['port']}")
    output_formula = '"{}"'.format(",".join(output_list)) if output_list else '"None"'

    # Build input list for output dropdowns (destination selection)
    input_list = []
    for d in DEVICES:
        for i in d["inputs"]:
            input_list.append(f"{d['name']}: {i['port']}")
    input_formula = '"{}"'.format(",".join(input_list)) if input_list else '"None"'

    # Add validations to device sheets
    for device in DEVICES:
        ws = wb[device["name"]]

        # Source dropdown for inputs (column D)
        if device["inputs"]:
            src_dv = DataValidation(type="list", formula1=output_formula, allow_blank=True)
            ws.add_data_validation(src_dv)
            src_dv.add(f'D7:D{6 + len(device["inputs"])}')

        # Destination dropdown for outputs (column H)
        if device["outputs"]:
            dst_dv = DataValidation(type="list", formula1=input_formula, allow_blank=True)
            ws.add_data_validation(dst_dv)
            dst_dv.add(f'H7:H{6 + len(device["outputs"])}')

    # ===== SAVE =====
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"AV Router created: {output_path}")
    return output_path

if __name__ == "__main__":
    create_av_router()
